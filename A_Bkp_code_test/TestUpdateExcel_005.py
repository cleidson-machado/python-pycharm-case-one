import openpyxl as ex
from Util import PlaceStorageStrings as Path

wb = ex.load_workbook(Path.output_folder + Path.report_002)

# DEFINE THE SHEET NAME TO WORK... THE SHEET SOURCE
# sh = wb['Sheet1']
sh = wb[wb.active.title]

# for col in sh.iter_rows(min_row=2, min_col=4, max_col=6):
#     for cell in sh['C']:
#         new_value = (['RED_PLUS'])
#         sh.append(new_value)

# for index, row in sh.iter_rows():
    # if sh.loc[index, 'EMAIL_ACT'] == 'RED':
    #     print('Passei aqui')

# JUST A COPY
for col in sh.iter_rows(min_row=2, min_col=4, max_col=6):
    for cell in sh['C']:
        if sh.loc['C'] == 'RED':
            print('%s: cell.value=%s' % (cell, cell.value))
