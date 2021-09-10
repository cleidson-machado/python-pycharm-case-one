import openpyxl as ex
from Util import PlaceStorageStrings as Path

myFileName = Path.output_folder + Path.report_002

wb = ex.load_workbook(filename=myFileName)

ws = wb['Sheet1']

newRowLocation = ws.max_row + 1

# ws.cell(column=1, row=newRowLocation, value="aha! a new entry at the end")
# wb.save(filename=myFileName)
# wb.close()

for index, row in ws.iterrows():
    if ws.loc[index, 'EMAIL_ACT'] == 'RED':
        print('Passei aqui')












# LOOK AGAIN!!
# https://stackoverflow.com/questions/36909977/update-row-values-where-certain-condition-is-met-in-pandas/36910033


# sh = wb[wb.active.title]
#
# for col in sh.iter_rows(min_row=2, min_col=4, max_col=6):
#     for cell in sh['C']:
#         new_value = (['RED_PLUS'])
#         sh.append(new_value)


# FROM HERE!!
# https://stackoverflow.com/questions/59167285/how-to-append-data-to-the-last-row-every-time-of-an-excel-file
