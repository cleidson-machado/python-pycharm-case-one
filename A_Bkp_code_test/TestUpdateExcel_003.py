import openpyxl as ex
from Util import PlaceStorageStrings as Path

wb = ex.load_workbook(Path.output_folder + Path.report_002)

# DEFINE THE SHEET NAME TO WORK... THE SHEET SOURCE
sh = wb['dados_teste']

# COUNT THE ROW'S
row = sh.max_row

# COUNT THE COLUMN'S
column = sh.max_column

for i in range(1, row):
    for cell in sh['C']:
        print(cell.value)
