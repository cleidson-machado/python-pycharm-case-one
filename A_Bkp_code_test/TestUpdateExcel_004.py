import openpyxl as ex
from Util import PlaceStorageStrings as Path

wb = ex.load_workbook(Path.output_folder + Path.report_002)

# DEFINE THE SHEET NAME TO WORK... THE SHEET SOURCE
# sh = wb['Sheet1']
sh = wb[wb.active.title]

# COUNT THE ROW'S
row = sh.max_row

# COUNT THE COLUMN'S
column = sh.max_column

for col in sh.iter_rows(min_row=2,min_col=4,max_col=6):
    for cell in sh['C']:
        print('%s: cell.value=%s' % (cell, cell.value))
        # new_value = "RED_PLUS"
        # sh.append(new_value)




# https://gist.github.com/armaandhir/455828396bf5c18256da














# https://stackoverflow.com/questions/54261748/skip-first-row-in-openpxyl/54262175
# https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.worksheet.html#openpyxl.worksheet.worksheet.Worksheet.iter_rows
# https://stackoverflow.com/questions/34754077/openpyxl-how-to-read-only-one-column-from-excel-file-in-python
# https://openpyxl.readthedocs.io/en/stable/usage.html#read-an-existing-workbook
