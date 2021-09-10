import openpyxl as ex
from Util import PlaceStorageStrings as Path

wb = ex.load_workbook(Path.output_folder + Path.report_002)

# DEFINE THE SHEET NAME TO WORK
sh = wb['dados_teste']

# COUNT THE ROW'S
row = sh.max_row

# COUNT THE COLUMN'S
column = sh.max_column

# HERE JUST GET THE SHEET NAME
sheets = wb.sheetnames

# print(sheets) # nome da planilha
# print(wb.active.title)  # nome da planilha ativa..

# HERE BASIC NAVIGATE ON SHEET VALUES..
for i in range(1, row + 1):
    for j in range(1, column + 1):
        print(sh.cell(i, j).value)

# HERE BASIC NAVIGATE ON SHEET VALUES.. WORK THE SAME WAY...
for i in range(1, row):
    for j in range(1, column):
        print(sh.cell(i, j).value)
